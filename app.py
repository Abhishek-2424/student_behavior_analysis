from flask import Flask, render_template, request, send_from_directory, url_for, jsonify
import os
import cv2
import numpy as np
from ultralytics import YOLO
import base64
from PIL import Image
import io
import torch
from datetime import datetime
import concurrent.futures
import queue
import threading
import pandas as pd
import time
import openpyxl
import tempfile
import shutil
from openpyxl.styles import Font, PatternFill

# Get the absolute path of the current directory
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# Initialize Flask app with explicit static folder path
app = Flask(__name__,
            static_url_path='',  # Empty string to serve from root
            static_folder='static')     # Directory containing static files

# Enable debug mode for development
app.config['DEBUG'] = True

# Print debug information
print(f"Base Directory: {BASE_DIR}")
print(f"Static Folder: {app.static_folder}")
print(f"Static URL Path: {app.static_url_path}")

# Define file paths using absolute paths
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'data', 'images')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'output')
EXCEL_FOLDER = os.path.join(BASE_DIR, 'excel_reports')

# Create all necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(OUTPUT_FOLDER, 'images'), exist_ok=True)
os.makedirs(os.path.join(OUTPUT_FOLDER, 'videos'), exist_ok=True)
os.makedirs(os.path.join(app.static_folder, 'js'), exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)

# Load the trained YOLO model
model = YOLO('models/best(4).pt')  # Path to your trained YOLOv8 model

# Behavior categories (ensure these match the model's class labels)
BEHAVIORS = ['hand raising', 'laptop', 'laughing', 'listening class', 'looking away', 'mobile phone', 'reading', 'sleeping', 'using laptop', 'using phone', 'writing']

# Categorize behaviors into Objects and Behaviors
BEHAVIOR_CATEGORIES = {
    'Objects': ['laptop', 'mobile phone'],
    'Behaviors': ['hand raising', 'laughing', 'listening class', 'looking away', 'reading', 'sleeping', 'using laptop', 'using phone', 'writing']
}

# Global dictionary to store behavior counts
behavior_tracking = {
    'timestamp': [],
    'total_students': [],
}
for behavior in BEHAVIORS:
    behavior_tracking[behavior] = []

# Add these global variables after other global variables
realtime_behavior_counts = {behavior: 0 for behavior in BEHAVIORS}
realtime_frame_max_counts = {behavior: 0 for behavior in BEHAVIORS}
realtime_video_writer = None
realtime_video_path = None
realtime_lock = threading.Lock()
is_recording = False

def export_to_excel():
    """Export behavior tracking data to Excel file with student counts per behavior"""
    try:
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'behavior_report_{timestamp}.xlsx'
        filepath = os.path.join(EXCEL_FOLDER, filename)
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Get the latest detection data
            latest_idx = -1 if behavior_tracking['timestamp'] else None
            
            # Create separate DataFrames for Objects and Behaviors
            objects_data = {'Detected Objects': [], 'Number of Objects': []}
            behaviors_data = {'Detected Behaviors': [], 'Number of Students': []}
            
            if latest_idx is not None:
                # Process Objects
                for obj in BEHAVIOR_CATEGORIES['Objects']:
                    count = behavior_tracking[obj][latest_idx] if behavior_tracking[obj] else 0
                    if count > 0:
                        objects_data['Detected Objects'].append(obj)
                        objects_data['Number of Objects'].append(int(count))
                
                # Process Behaviors
                for behavior in BEHAVIOR_CATEGORIES['Behaviors']:
                    count = behavior_tracking[behavior][latest_idx] if behavior_tracking[behavior] else 0
                    if count > 0:
                        behaviors_data['Detected Behaviors'].append(behavior)
                        behaviors_data['Number of Students'].append(int(count))
            
            # Create DataFrames
            objects_df = pd.DataFrame(objects_data)
            behaviors_df = pd.DataFrame(behaviors_data)
            
            # Write Objects section
            start_row = 0
            objects_df.to_excel(writer, sheet_name='Detection Summary', startrow=start_row, index=False)
            
            # Calculate total objects
            total_objects = sum(objects_data['Number of Objects'])
            
            # Write Behaviors section
            start_row = len(objects_data['Detected Objects']) + 3  # Add space between sections
            behaviors_df.to_excel(writer, sheet_name='Detection Summary', startrow=start_row, index=False)
            
            # Calculate total students
            total_students = sum(behaviors_data['Number of Students'])
            
            # Get the worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Detection Summary']
            
            # Format headers
            header_font = openpyxl.styles.Font(bold=True)
            
            # Format Objects section
            worksheet['A1'] = 'Detected Objects'
            worksheet['B1'] = 'Number of Objects'
            worksheet.cell(row=1, column=1).font = header_font
            worksheet.cell(row=1, column=2).font = header_font
            
            # Format Behaviors section
            worksheet.cell(row=start_row + 1, column=1, value='Detected Behaviors')
            worksheet.cell(row=start_row + 1, column=2, value='Number of Students')
            worksheet.cell(row=start_row + 1, column=1).font = header_font
            worksheet.cell(row=start_row + 1, column=2).font = header_font
            
            # Add section totals
            objects_total_row = len(objects_data['Detected Objects']) + 2
            worksheet[f'A{objects_total_row}'] = 'Total Objects'
            worksheet[f'B{objects_total_row}'] = total_objects
            worksheet[f'A{objects_total_row}'].font = header_font
            
            behaviors_total_row = start_row + len(behaviors_data['Detected Behaviors']) + 2
            worksheet[f'A{behaviors_total_row}'] = 'Total Students'
            worksheet[f'B{behaviors_total_row}'] = total_students
            worksheet[f'A{behaviors_total_row}'].font = header_font
            
            # Add timestamp at the bottom
            timestamp_row = behaviors_total_row + 2
            worksheet[f'A{timestamp_row}'] = 'Detection Time'
            worksheet[f'B{timestamp_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            worksheet[f'A{timestamp_row}'].font = header_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        return filename
    except Exception as e:
        print(f"Error in export_to_excel: {str(e)}")
        return None

def update_behavior_tracking(behaviors_dict):
    """Update behavior tracking with new detection results"""
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    behavior_tracking['timestamp'].append(current_time)
    
    # Count total number of students (sum of all behaviors)
    total_students = sum(behaviors_dict.values())
    behavior_tracking['total_students'].append(total_students)
    
    # Update individual behavior counts
    for behavior in BEHAVIORS:
        behavior_tracking[behavior].append(behaviors_dict.get(behavior, 0))

# Route for handling form submission
@app.route('/')
def index():
    return render_template('index.html')

# Upload and process image/video
@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return 'No file part', 400
        
        file = request.files['file']
        if file.filename == '':
            return 'No selected file', 400

        # Create a safe filename
        filename = file.filename.replace(' ', '_')
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        # Save the uploaded file
        file.save(file_path)

        # Check if the file is an image or a video
        if filename.lower().endswith(('.mp4', '.avi', '.mov')):
            output_video_path = os.path.join(OUTPUT_FOLDER, 'videos', filename)
            # Get behavior counts from video processing
            behaviors = process_video(file_path, output_video_path)
            output_video_url = url_for('static_file', folder='videos', filename=filename)
            
            # Export results to Excel
            excel_filename = export_to_excel()
            excel_download_url = url_for('download_excel', filename=excel_filename)
            
            return render_template('index.html', 
                                video_url=output_video_url, 
                                download_url=output_video_url, 
                                filename=filename,
                                behaviors=behaviors,
                                excel_url=excel_download_url,
                                show_processed=True)
        
        elif filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif')):
            output_image_path = os.path.join(OUTPUT_FOLDER, 'images', filename)
            behaviors, output_image_url = process_image(file_path, output_image_path)
            
            # Export results to Excel
            excel_filename = export_to_excel()
            excel_download_url = url_for('download_excel', filename=excel_filename)
            
            return render_template('index.html', 
                                image_url=output_image_url, 
                                download_url=output_image_url, 
                                filename=filename, 
                                behaviors=behaviors,
                                excel_url=excel_download_url,
                                show_processed=True)
        
        else:
            return 'Unsupported file type. Please upload an image or video file.', 400

    except Exception as e:
        print(f"Error processing upload: {str(e)}")
        return f'Error processing file: {str(e)}', 500

# Process image with YOLO and count behaviors
def process_image(image_path, output_path):
    img = cv2.imread(image_path)
    results = model(img)  # Perform inference

    # Get class names and their corresponding counts
    behavior_counts = {behavior: 0 for behavior in BEHAVIORS}

    # For each detection, check the class and update the count
    for result in results:
        for detection in result.boxes:
            class_id = int(detection.cls[0].item())
            if 0 <= class_id < len(BEHAVIORS):
                behavior = BEHAVIORS[class_id]
                behavior_counts[behavior] += 1

    # Update behavior tracking before plotting
    update_behavior_tracking(behavior_counts)

    # Plot the bounding boxes on the image
    output_image = results[0].plot()
    cv2.imwrite(output_path, output_image)

    # Return behavior counts and output image URL
    output_image_url = url_for('static_file', folder='images', filename=os.path.basename(output_path))
    return behavior_counts, output_image_url

# Process video with YOLO
def process_video(video_path, output_video_path):
    cap = cv2.VideoCapture(video_path)
    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    
    # Initialize behavior counters with maximum counts
    behavior_counts = {behavior: 0 for behavior in BEHAVIORS}
    frame_max_counts = {behavior: 0 for behavior in BEHAVIORS}
    processed_frames = 0
    
    # Use H.264 codec for better web compatibility
    if os.name == 'nt':  # Windows
        fourcc = cv2.VideoWriter_fourcc(*'avc1')
    else:  # Linux/Mac
        fourcc = cv2.VideoWriter_fourcc(*'avc1')
    
    temp_output_path = output_video_path.replace('.mp4', '_temp.mp4')
    out = cv2.VideoWriter(temp_output_path, fourcc, fps, (frame_width, frame_height))

    # Brightness adjustment parameters
    alpha = 1.4  # Contrast control (1.0-3.0)
    beta = 25    # Brightness control (0-100)

    try:
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
                
            # Process frame with YOLO
            results = model(frame)
            
            # Update behavior counts for current frame
            frame_behaviors = {behavior: 0 for behavior in BEHAVIORS}
            for result in results:
                for detection in result.boxes:
                    class_id = int(detection.cls[0].item())
                    if 0 <= class_id < len(BEHAVIORS):
                        behavior = BEHAVIORS[class_id]
                        frame_behaviors[behavior] += 1
            
            # Update maximum counts for each behavior
            for behavior, count in frame_behaviors.items():
                frame_max_counts[behavior] = max(frame_max_counts[behavior], count)
                
                # For objects (laptop and mobile phone), keep the maximum count seen
                if behavior in BEHAVIOR_CATEGORIES['Objects']:
                    behavior_counts[behavior] = frame_max_counts[behavior]
                # For behaviors, if it's detected in this frame, mark it as detected
                else:
                    if count > 0:
                        behavior_counts[behavior] = 1
            
            # Update tracking for each frame
            update_behavior_tracking(frame_behaviors)
            
            # Draw detections on frame
            frame_output = results[0].plot()
            
            # Adjust brightness and contrast
            frame_output = cv2.convertScaleAbs(frame_output, alpha=alpha, beta=beta)
            
            out.write(frame_output)
            processed_frames += 1
            
    finally:
        # Release resources
        cap.release()
        out.release()

    # Convert video to web-compatible format using FFmpeg if available
    try:
        import subprocess
        ffmpeg_cmd = [
            'ffmpeg', '-i', temp_output_path,
            '-vcodec', 'libx264',
            '-acodec', 'aac',
            '-movflags', '+faststart',
            '-y',  # Overwrite output file if it exists
            output_video_path
        ]
        subprocess.run(ffmpeg_cmd, check=True)
        os.remove(temp_output_path)  # Remove temporary file
    except Exception as e:
        print(f"FFmpeg conversion failed: {str(e)}")
        # If FFmpeg fails, just use the original file
        if os.path.exists(temp_output_path):
            os.replace(temp_output_path, output_video_path)
    
    return behavior_counts

# Serve the processed static files
@app.route('/static/<folder>/<filename>')
def static_file(folder, filename):
    directory = os.path.join(OUTPUT_FOLDER, folder)
    if folder == 'videos':
        response = send_from_directory(directory, filename, as_attachment=False)
        # Set correct MIME type for MP4 videos
        response.headers['Content-Type'] = 'video/mp4'
        # Add headers to help with video streaming
        response.headers['Accept-Ranges'] = 'bytes'
        return response
    return send_from_directory(directory, filename, as_attachment=True)

# Process real-time video frames
@app.route('/start_realtime', methods=['POST'])
def start_realtime():
    global realtime_video_writer, realtime_video_path, is_recording
    global realtime_behavior_counts, realtime_frame_max_counts
    
    try:
        # Reset counters
        realtime_behavior_counts = {behavior: 0 for behavior in BEHAVIORS}
        realtime_frame_max_counts = {behavior: 0 for behavior in BEHAVIORS}
        
        # Create temporary video file
        temp_dir = os.path.join(OUTPUT_FOLDER, 'videos')
        os.makedirs(temp_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        realtime_video_path = os.path.join(temp_dir, f'realtime_{timestamp}.mp4')
        
        # Initialize video writer (we'll set the dimensions when we get the first frame)
        is_recording = True
        
        return jsonify({'status': 'success', 'message': 'Real-time session started'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/stop_realtime', methods=['POST'])
def stop_realtime():
    global realtime_video_writer, is_recording
    
    try:
        is_recording = False
        if realtime_video_writer:
            realtime_video_writer.release()
            realtime_video_writer = None

        # Get the session data from the request
        data = request.get_json()
        session_data = data.get('sessionData', {})
        detected_objects = session_data.get('detectedObjects', {})
        detected_behaviors = session_data.get('detectedBehaviors', {})
        start_time = session_data.get('startTime')
        end_time = session_data.get('timestamp')
        total_frames = session_data.get('frames', 0)
        
        # Create Excel file with final data
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'realtime_detection_{timestamp}.xlsx'
        excel_path = os.path.join(EXCEL_FOLDER, excel_filename)
        
        # Create workbook with custom styles
        workbook = openpyxl.Workbook()
        
        # Define styles
        header_style = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        # Create Objects sheet
        objects_sheet = workbook.active
        objects_sheet.title = 'Detected Objects'
        
        # Set headers with styling
        headers = ['Object Type', 'Count', 'Detection Time']
        for col, header in enumerate(headers, 1):
            cell = objects_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        # Add detected objects data
        row = 2
        for obj, count in detected_objects.items():
            if count > 0:
                objects_sheet.cell(row=row, column=1, value=obj)
                objects_sheet.cell(row=row, column=2, value=count)
                objects_sheet.cell(row=row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                row += 1
        
        # Create Behaviors sheet
        behaviors_sheet = workbook.create_sheet('Detected Behaviors')
        
        # Set headers with styling
        for col, header in enumerate(['Behavior Type', 'Count', 'Detection Time'], 1):
            cell = behaviors_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        # Add detected behaviors data
        row = 2
        for behavior, count in detected_behaviors.items():
            if count > 0:
                behaviors_sheet.cell(row=row, column=1, value=behavior)
                behaviors_sheet.cell(row=row, column=2, value=count)
                behaviors_sheet.cell(row=row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                row += 1
        
        # Create Session Info sheet
        info_sheet = workbook.create_sheet('Session Information')
        
        # Set headers with styling
        info_sheet.cell(row=1, column=1, value='Metric')
        info_sheet.cell(row=1, column=2, value='Value')
        info_sheet.cell(row=1, column=1).font = header_style
        info_sheet.cell(row=1, column=2).font = header_style
        info_sheet.cell(row=1, column=1).fill = header_fill
        info_sheet.cell(row=1, column=2).fill = header_fill
        
        # Add session information
        session_info = [
            ('Session Start Time', start_time),
            ('Session End Time', end_time),
            ('Total Frames Processed', total_frames),
            ('Total Objects Detected', sum(detected_objects.values())),
            ('Total Behaviors Detected', len([b for b, c in detected_behaviors.items() if c > 0])),
            ('Average FPS', round(total_frames / ((datetime.fromisoformat(end_time) - datetime.fromisoformat(start_time)).total_seconds()), 2) if start_time and end_time else 'N/A')
        ]
        
        for row, (metric, value) in enumerate(session_info, 2):
            info_sheet.cell(row=row, column=1, value=metric)
            info_sheet.cell(row=row, column=2, value=value)
        
        # Auto-adjust column widths for all sheets
        for sheet in workbook:
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(excel_path)
        
        return jsonify({
            'status': 'success',
            'message': 'Real-time session stopped',
            'excel_url': url_for('download_excel', filename=excel_filename)
        })
        
    except Exception as e:
        print(f"Error stopping realtime: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error stopping session: {str(e)}'
        }), 500

# Process real-time video frames
@app.route('/process_frame', methods=['POST'])
def process_frame():
    global realtime_video_writer, realtime_behavior_counts, realtime_frame_max_counts, is_recording
    
    try:
        # Get the frame data from the request
        data = request.get_json()
        frame_data = data['frame'].split(',')[1]  # Remove the data URL prefix
        
        # Convert base64 to image
        frame_bytes = base64.b64decode(frame_data)
        frame_arr = np.frombuffer(frame_bytes, dtype=np.uint8)
        frame = cv2.imdecode(frame_arr, cv2.IMREAD_COLOR)
        
        # Initialize video writer if not already done
        if is_recording and realtime_video_writer is None and frame is not None:
            height, width = frame.shape[:2]
            fourcc = cv2.VideoWriter_fourcc(*'avc1')
            realtime_video_writer = cv2.VideoWriter(realtime_video_path, fourcc, 20.0, (width, height))
        
        # Process the frame with YOLO
        results = model(frame)
        
        # Update behavior counts for current frame
        frame_behaviors = {behavior: 0 for behavior in BEHAVIORS}
        
        for result in results:
            for detection in result.boxes:
                class_id = int(detection.cls[0].item())
                if 0 <= class_id < len(BEHAVIORS):
                    behavior = BEHAVIORS[class_id]
                    frame_behaviors[behavior] += 1
        
        # Update maximum counts and behavior detection with lock
        with realtime_lock:
            for behavior, count in frame_behaviors.items():
                realtime_frame_max_counts[behavior] = max(realtime_frame_max_counts[behavior], count)
                
                # For objects, keep the maximum count seen
                if behavior in BEHAVIOR_CATEGORIES['Objects']:
                    realtime_behavior_counts[behavior] = realtime_frame_max_counts[behavior]
                # For behaviors, if detected, mark as present
                else:
                    if count > 0:
                        realtime_behavior_counts[behavior] = 1
        
        # Update behavior tracking
        update_behavior_tracking(frame_behaviors)
        
        # Get the processed frame with visualizations
        processed_frame = results[0].plot()
        
        # Write frame to video if recording
        if is_recording and realtime_video_writer is not None:
            realtime_video_writer.write(processed_frame)
        
        # Convert the processed frame to base64
        _, buffer = cv2.imencode('.jpg', processed_frame)
        processed_frame_data = base64.b64encode(buffer).decode('utf-8')
        
        # Export Excel report periodically (every 30 frames)
        if sum(frame_behaviors.values()) > 0:
            excel_filename = export_to_excel()
        
        return jsonify({
            'behaviors': realtime_behavior_counts,
            'frame_data': processed_frame_data
        })
    
    except Exception as e:
        print(f"Error processing frame: {str(e)}")
        return jsonify({'error': str(e)}), 500

def process_frame_yolo(frame):
    """Process a single frame with YOLO model"""
    # Convert frame to RGB if needed
    if len(frame.shape) == 2:  # If grayscale
        frame = cv2.cvtColor(frame, cv2.COLOR_GRAY2RGB)
    elif frame.shape[2] == 4:  # If RGBA
        frame = cv2.cvtColor(frame, cv2.COLOR_RGBA2RGB)
    
    # Inference
    results = model(frame)
    
    # Get detections
    pred = results.pred[0]
    behaviors = {}
    
    if len(pred):
        for *box, conf, cls in pred:
            behavior = results.names[int(cls)]
            behaviors[behavior] = behaviors.get(behavior, 0) + 1
            
            # Draw bounding box
            x1, y1, x2, y2 = map(int, box)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(frame, f"{behavior} {conf:.2f}", (x1, y1 - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
    
    return frame, behaviors

def process_video_optimized(video_path, output_path):
    """Process video with optimized performance"""
    cap = cv2.VideoCapture(video_path)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = int(cap.get(cv2.CAP_PROP_FPS))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    # Create video writer
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    
    # Initialize behavior counters
    total_behaviors = {}
    processed_frames = 0
    
    # Create a queue for frames and results
    frame_queue = queue.Queue(maxsize=30)  # Limit queue size
    result_queue = queue.Queue()
    
    def process_frame_worker():
        while True:
            frame_data = frame_queue.get()
            if frame_data is None:  # Sentinel value to stop the thread
                break
            frame_number, frame = frame_data
            processed_frame, behaviors = process_frame_yolo(frame)
            result_queue.put((frame_number, processed_frame, behaviors))
            frame_queue.task_done()
    
    # Start worker threads
    num_workers = min(4, os.cpu_count() or 1)  # Use up to 4 threads
    workers = []
    for _ in range(num_workers):
        worker = threading.Thread(target=process_frame_worker)
        worker.daemon = True
        worker.start()
        workers.append(worker)
    
    frame_number = 0
    results_buffer = {}
    next_frame_to_write = 0
    
    try:
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            
            # Skip frames if processing is falling behind
            if frame_number % 2 != 0:  # Process every other frame
                frame_number += 1
                continue
            
            # Add frame to queue
            frame_queue.put((frame_number, frame))
            
            # Get processed results
            while not result_queue.empty():
                idx, processed_frame, behaviors = result_queue.get()
                results_buffer[idx] = (processed_frame, behaviors)
                
                # Write frames in order
                while next_frame_to_write in results_buffer:
                    frame_to_write, frame_behaviors = results_buffer.pop(next_frame_to_write)
                    out.write(frame_to_write)
                    # Update total behaviors
                    for behavior, count in frame_behaviors.items():
                        total_behaviors[behavior] = total_behaviors.get(behavior, 0) + count
                    next_frame_to_write += 2
                    processed_frames += 1
            
            frame_number += 1
    
    finally:
        # Signal workers to stop
        for _ in workers:
            frame_queue.put(None)
        
        # Wait for workers to finish
        for worker in workers:
            worker.join()
        
        # Process remaining results
        while not result_queue.empty():
            idx, processed_frame, behaviors = result_queue.get()
            results_buffer[idx] = (processed_frame, behaviors)
        
        # Write remaining frames in order
        for idx in sorted(results_buffer.keys()):
            frame_to_write, frame_behaviors = results_buffer[idx]
            out.write(frame_to_write)
            for behavior, count in frame_behaviors.items():
                total_behaviors[behavior] = total_behaviors.get(behavior, 0) + count
            processed_frames += 1
        
        # Release resources
        cap.release()
        out.release()
    
    # Calculate average behaviors per processed frame
    for behavior in total_behaviors:
        total_behaviors[behavior] = int(total_behaviors[behavior] / processed_frames * total_frames)
    
    return total_behaviors

# Serve static files directly
@app.route('/static/js/<path:filename>')
def serve_static(filename):
    return send_from_directory(os.path.join(app.static_folder, 'js'), filename)

# Serve realtime.js directly
@app.route('/static/js/realtime.js')
def serve_realtime_js():
    return send_from_directory(
        os.path.join(app.static_folder, 'js'),
        'realtime.js',
        mimetype='application/javascript'
    )

@app.route('/download_excel/<filename>')
def download_excel(filename):
    """Endpoint to download Excel reports"""
    return send_from_directory(EXCEL_FOLDER, filename, as_attachment=True)

@app.route('/get_current_results', methods=['POST'])
def get_current_results():
    try:
        data = request.get_json()
        session_data = data.get('sessionData', {})
        
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'realtime_detection_{timestamp}.xlsx'
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        
        # Create a new workbook
        workbook = openpyxl.Workbook()
        
        # Create Objects sheet
        objects_sheet = workbook.active
        objects_sheet.title = 'Detected Objects'
        objects_sheet['A1'] = 'Object'
        objects_sheet['B1'] = 'Total Count'
        objects_sheet['C1'] = 'Average per Frame'
        
        row = 2
        for obj, count in session_data.get('detectedObjects', {}).items():
            objects_sheet[f'A{row}'] = obj
            objects_sheet[f'B{row}'] = count
            objects_sheet[f'C{row}'] = session_data.get('averageBehaviors', {}).get(obj, 0)
            row += 1
        
        # Create Behaviors sheet
        behaviors_sheet = workbook.create_sheet('Detected Behaviors')
        behaviors_sheet['A1'] = 'Behavior'
        behaviors_sheet['B1'] = 'Total Count'
        behaviors_sheet['C1'] = 'Average per Frame'
        
        row = 2
        for behavior, count in session_data.get('detectedBehaviors', {}).items():
            behaviors_sheet[f'A{row}'] = behavior
            behaviors_sheet[f'B{row}'] = count
            behaviors_sheet[f'C{row}'] = session_data.get('averageBehaviors', {}).get(behavior, 0)
            row += 1
        
        # Create Session Info sheet
        info_sheet = workbook.create_sheet('Session Information')
        info_sheet['A1'] = 'Metric'
        info_sheet['B1'] = 'Value'
        
        info_sheet['A2'] = 'Start Time'
        info_sheet['B2'] = session_data.get('timestamp', '')
        info_sheet['A3'] = 'Total Frames'
        info_sheet['B3'] = session_data.get('frames', 0)
        info_sheet['A4'] = 'Duration (seconds)'
        info_sheet['B4'] = (datetime.now() - datetime.fromisoformat(session_data.get('timestamp', datetime.now().isoformat()))).total_seconds()
        
        # Apply styling
        for sheet in workbook:
            # Style headers
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            
            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(excel_path)
        
        # Return the URL for the Excel file
        excel_url = url_for('static', filename=f'uploads/{excel_filename}')
        return jsonify({
            'status': 'success',
            'excel_url': excel_url
        })
        
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to generate Excel file'
        }), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)  # Enable debug mode for better error messages
